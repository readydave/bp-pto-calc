from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

import openpyxl

from .calc import generate_default_holidays
from .models import (
    SCHEMA_VERSION,
    EntrySource,
    EntryType,
    HolidayEntry,
    PlannedEntry,
    ProjectionResult,
    ProjectionRowType,
    PtoScenario,
    ScenarioPolicy,
)


def save_scenario(path: str | Path, scenario: PtoScenario) -> None:
    payload = scenario_to_payload(scenario)
    Path(path).write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_scenario(path: str | Path) -> PtoScenario:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") == SCHEMA_VERSION:
        return scenario_from_payload(payload)
    return legacy_scenario_from_payload(payload, Path(path).stem)


def scenario_to_payload(scenario: PtoScenario) -> dict:
    return {
        "schema_version": SCHEMA_VERSION,
        "scenario": {
            "name": scenario.name,
            "year": scenario.year,
            "regular_balance": scenario.regular_balance,
            "accrual_per_period": scenario.accrual_per_period,
            "float_balance": scenario.float_balance,
            "last_pay_date": scenario.last_pay_date.isoformat() if scenario.last_pay_date else None,
        },
        "policy": {
            "regular_pto_cap": scenario.policy.regular_pto_cap,
            "float_award_per_quarter": scenario.policy.float_award_per_quarter,
            "holiday_hours": scenario.policy.holiday_hours,
        },
        "planned_entries": [
            {
                "date": planned_entry.date.isoformat(),
                "hours": planned_entry.hours,
                "entry_type": planned_entry.entry_type.value,
                "note": planned_entry.note,
                "source": planned_entry.source.value,
            }
            for planned_entry in scenario.planned_entries
        ],
        "holiday_overrides": [
            {
                "date": holiday.date.isoformat(),
                "name": holiday.name,
                "hours": holiday.hours,
                "enabled": holiday.enabled,
            }
            for holiday in scenario.holidays
        ],
    }


def scenario_from_payload(payload: dict) -> PtoScenario:
    scenario_data = payload["scenario"]
    policy_data = payload["policy"]
    planned_entries = [
        PlannedEntry(
            date=date.fromisoformat(entry["date"]),
            hours=float(entry["hours"]),
            entry_type=EntryType(entry["entry_type"]),
            note=entry.get("note", ""),
            source=EntrySource(entry.get("source", EntrySource.MANUAL.value)),
        )
        for entry in payload.get("planned_entries", [])
    ]
    holidays = [
        HolidayEntry(
            date=date.fromisoformat(holiday["date"]),
            name=holiday["name"],
            hours=float(holiday["hours"]),
            enabled=bool(holiday.get("enabled", True)),
        )
        for holiday in payload.get("holiday_overrides", [])
    ]
    return PtoScenario(
        name=scenario_data.get("name", "Untitled Scenario"),
        year=int(scenario_data["year"]),
        regular_balance=float(scenario_data.get("regular_balance", 0.0)),
        accrual_per_period=float(scenario_data.get("accrual_per_period", 0.0)),
        float_balance=float(scenario_data.get("float_balance", 0.0)),
        last_pay_date=date.fromisoformat(scenario_data["last_pay_date"]) if scenario_data.get("last_pay_date") else None,
        policy=ScenarioPolicy(
            regular_pto_cap=float(policy_data.get("regular_pto_cap", 120.0)),
            float_award_per_quarter=float(policy_data.get("float_award_per_quarter", 12.0)),
            holiday_hours=float(policy_data.get("holiday_hours", 8.0)),
        ),
        planned_entries=planned_entries,
        holidays=holidays,
    )


def legacy_scenario_from_payload(payload: dict, scenario_name: str = "Imported Scenario") -> PtoScenario:
    last_pay_date = date.fromisoformat(payload["last_pay_date"]) if payload.get("last_pay_date") else None
    year = last_pay_date.year if last_pay_date else date.today().year
    policy = ScenarioPolicy()
    planned_entries = [
        PlannedEntry(
            date=date.fromisoformat(entry_date),
            hours=float(hours),
            entry_type=EntryType(entry_type),
            source=EntrySource.MANUAL,
        )
        for entry_date, hours, entry_type in payload.get("planned_pto", [])
    ]
    return PtoScenario(
        name=scenario_name or "Imported Scenario",
        year=year,
        regular_balance=float(payload.get("regular_balance", 0.0)),
        accrual_per_period=float(payload.get("accrual", 0.0)),
        float_balance=float(payload.get("float_balance", 0.0)),
        last_pay_date=last_pay_date,
        policy=policy,
        planned_entries=planned_entries,
        holidays=generate_default_holidays(year, policy.holiday_hours),
    )


def export_projection_to_csv(path: str | Path, result: ProjectionResult) -> None:
    with Path(path).open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "Date",
                "Type",
                "Label",
                "Regular Start",
                "Float Start",
                "Regular Change",
                "Float Change",
                "Regular Used",
                "Float Used",
                "Regular End",
                "Float End",
                "Note",
            ]
        )
        for row in result.rows:
            writer.writerow(_projection_row_csv_values(row))


def export_projection_to_excel(path: str | Path, result: ProjectionResult) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"PTO Planner {result.year}"

    headers = [
        "Date",
        "Type",
        "Label",
        "Regular Start",
        "Float Start",
        "Regular Change",
        "Float Change",
        "Regular Used",
        "Float Used",
        "Regular End",
        "Float End",
        "Note",
    ]
    worksheet.append(headers)

    header_font = openpyxl.styles.Font(bold=True)
    fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="DCEAF7")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    for row in result.rows:
        worksheet.append(_projection_row_values(row))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    date_format = "yyyy-mm-dd"
    number_format = "0.0000"
    for excel_row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        excel_row[0].number_format = date_format
        for cell in excel_row[3:11]:
            cell.number_format = number_format

    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = max_length

    workbook.save(path)


def _projection_row_values(row) -> list:
    return [
        row.date,
        _projection_type_label(row.row_type),
        row.label,
        row.regular_start,
        row.float_start,
        row.regular_change,
        row.float_change,
        row.regular_used,
        row.float_used,
        row.regular_end,
        row.float_end,
        row.note,
    ]


def _projection_row_csv_values(row) -> list[str]:
    values = _projection_row_values(row)
    return [
        values[0].isoformat(),
        values[1],
        values[2],
        f"{values[3]:.4f}",
        f"{values[4]:.4f}",
        f"{values[5]:.4f}",
        f"{values[6]:.4f}",
        f"{values[7]:.4f}",
        f"{values[8]:.4f}",
        f"{values[9]:.4f}",
        f"{values[10]:.4f}",
        values[11],
    ]


def _projection_type_label(row_type: ProjectionRowType) -> str:
    labels = {
        ProjectionRowType.PAYDAY: "Payday",
        ProjectionRowType.FLOAT_AWARD: "Float Award",
        ProjectionRowType.PLANNED_PTO: "Planned PTO",
        ProjectionRowType.YEAR_END: "Year End",
    }
    return labels[row_type]
